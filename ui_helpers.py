"""
BADER Derneği - UI Helper Fonksiyonları
Ortak kullanılan UI yardımcı fonksiyonları
"""

from PyQt5.QtWidgets import (QComboBox, QCompleter, QTableWidget, QHeaderView, 
                             QMenu, QAction, QCheckBox, QWidgetAction, QFrame,
                             QVBoxLayout, QHBoxLayout, QPushButton, QLabel, QDialog,
                             QListWidget, QListWidgetItem, QMessageBox, QFileDialog)
from PyQt5.QtCore import Qt, QSettings
from PyQt5.QtGui import QCursor


def make_searchable_combobox(combobox: QComboBox):
    """
    ComboBox'ı aranabilir hale getirir
    
    Args:
        combobox: Aranabilir yapılacak QComboBox
    """
    # Düzenlenebilir yap
    combobox.setEditable(True)
    
    # Insert policy: kullanıcı yazarsa eklenmesin
    combobox.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
    
    # Completer ekle
    completer = QCompleter(combobox.model())
    completer.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
    completer.setCompletionMode(QCompleter.CompletionMode.PopupCompletion)
    completer.setFilterMode(Qt.MatchFlag.MatchContains)
    
    combobox.setCompleter(completer)
    
    # Stil düzeltmeleri
    combobox.lineEdit().setPlaceholderText("Arama veya seçim yapın...")
    
    return combobox


def create_searchable_combobox(items: list = None) -> QComboBox:
    """
    Yeni aranabilir ComboBox oluşturur
    
    Args:
        items: ComboBox'a eklenecek öğeler listesi
        
    Returns:
        Aranabilir QComboBox
    """
    combobox = QComboBox()
    
    if items:
        combobox.addItems(items)
    
    return make_searchable_combobox(combobox)


def add_separator_to_layout(layout, height: int = 1):
    """Layout'a ayırıcı çizgi ekler"""
    from PyQt5.QtWidgets import QFrame
    
    line = QFrame()
    line.setFrameShape(QFrame.Shape.HLine)
    line.setFrameShadow(QFrame.Shadow.Sunken)
    line.setMaximumHeight(height)
    line.setStyleSheet("background-color: #E0E0E0;")
    
    layout.addWidget(line)
    
    return line


def format_currency(amount: float, currency: str = "₺") -> str:
    """
    Para birimini formatlar
    
    Args:
        amount: Tutar
        currency: Para birimi sembolü
        
    Returns:
        Formatlanmış string
    """
    return f"{amount:,.2f} {currency}".replace(",", ".")


def format_date(date_str: str) -> str:
    """
    Tarih formatını düzenler
    
    Args:
        date_str: ISO format tarih (YYYY-MM-DD)
        
    Returns:
        Formatlanmış tarih (DD.MM.YYYY)
    """
    try:
        from datetime import datetime
        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
        return date_obj.strftime("%d.%m.%Y")
    except:
        return date_str


def export_table_to_excel(table, default_filename: str, parent_widget=None):
    """
    QTableWidget içeriğini Excel dosyasına export eder
    
    Args:
        table: QTableWidget nesnesi
        default_filename: Varsayılan dosya adı (uzantısız)
        parent_widget: Dialog için parent widget
        
    Returns:
        bool: Başarılı ise True
    """
    from PyQt5.QtWidgets import QFileDialog
    from qfluentwidgets import MessageBox
    from datetime import datetime
    
    # Dosya kaydet dialogu
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    suggested_name = f"{default_filename}_{timestamp}.xlsx"
    
    file_path, selected_filter = QFileDialog.getSaveFileName(
        parent_widget,
        "Excel'e Aktar",
        suggested_name,
        "Excel Dosyası (*.xlsx);;CSV Dosyası (*.csv);;Tüm Dosyalar (*)"
    )
    
    if not file_path:
        return False
    
    try:
        # Excel formatı için openpyxl dene
        if file_path.endswith('.xlsx') or 'Excel' in selected_filter:
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils import get_column_letter
                
                wb = Workbook()
                ws = wb.active
                ws.title = default_filename[:31]  # Max 31 karakter
                
                # Stil tanımları
                header_font = Font(bold=True, color="FFFFFF", size=11)
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Başlık satırı
                for col in range(table.columnCount()):
                    if table.isColumnHidden(col):
                        continue
                    header = table.horizontalHeaderItem(col)
                    cell = ws.cell(row=1, column=col+1, value=header.text() if header else f"Sütun {col+1}")
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border
                
                # Veri satırları
                row_num = 2
                for row in range(table.rowCount()):
                    if table.isRowHidden(row):
                        continue
                    
                    for col in range(table.columnCount()):
                        if table.isColumnHidden(col):
                            continue
                        item = table.item(row, col)
                        value = item.text() if item else ""
                        
                        # Sayısal değerleri dönüştür
                        try:
                            # Para birimi temizle ve sayıya çevir
                            clean_val = value.replace("₺", "").replace("TL", "").replace(".", "").replace(",", ".").strip()
                            if clean_val and clean_val.replace(".", "").replace("-", "").isdigit():
                                value = float(clean_val)
                        except:
                            pass
                        
                        cell = ws.cell(row=row_num, column=col+1, value=value)
                        cell.border = thin_border
                    
                    row_num += 1
                
                # Sütun genişliklerini ayarla
                for col in range(table.columnCount()):
                    if table.isColumnHidden(col):
                        continue
                    column_letter = get_column_letter(col + 1)
                    # Max genişliği bul
                    max_length = 0
                    for row in range(1, row_num):
                        cell = ws.cell(row=row, column=col+1)
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
                
                # Freeze panes (başlık satırını dondur)
                ws.freeze_panes = 'A2'
                
                # Kaydet
                if not file_path.endswith('.xlsx'):
                    file_path += '.xlsx'
                wb.save(file_path)
                
                MessageBox("Başarılı", f"Veriler Excel dosyasına aktarıldı:\n{file_path}", parent_widget).show()
                return True
                
            except ImportError:
                # openpyxl yoksa CSV'ye düş
                if not file_path.endswith('.csv'):
                    file_path = file_path.replace('.xlsx', '.csv')
        
        # CSV formatı
        import csv
        if not file_path.endswith('.csv'):
            file_path += '.csv'
            
        with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f, delimiter=';')
            
            # Başlık satırı
            headers = []
            for col in range(table.columnCount()):
                if table.isColumnHidden(col):
                    continue
                header = table.horizontalHeaderItem(col)
                headers.append(header.text() if header else f"Sütun {col+1}")
            writer.writerow(headers)
            
            # Veri satırları
            for row in range(table.rowCount()):
                if table.isRowHidden(row):
                    continue
                    
                row_data = []
                for col in range(table.columnCount()):
                    if table.isColumnHidden(col):
                        continue
                    item = table.item(row, col)
                    row_data.append(item.text() if item else "")
                writer.writerow(row_data)
        
        MessageBox("Başarılı", f"Veriler CSV dosyasına aktarıldı:\n{file_path}", parent_widget).show()
        return True
        
    except Exception as e:
        MessageBox("Hata", f"Export sırasında hata oluştu:\n{str(e)}", parent_widget).show()
        return False


def setup_resizable_table(table: QTableWidget, table_id: str = None, stretch_column: int = None):
    """
    Tabloyu responsive ve sütun genişlikleri ayarlanabilir yapar
    
    Args:
        table: QTableWidget nesnesi
        table_id: Ayarları kaydetmek için benzersiz ID
        stretch_column: Esnek genişlikli sütun indeksi (None ise son sütun)
    """
    header = table.horizontalHeader()
    
    # Tüm sütunları interaktif (hareket ettirilebilir) yap
    header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
    
    # Sütunları sürükleyerek taşınabilir yap
    header.setSectionsMovable(True)
    
    # Çift tıklama ile otomatik boyutlandır
    header.setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
    
    # Bir sütunu esnek yap (pencere genişliğine göre uzar)
    if stretch_column is not None:
        header.setSectionResizeMode(stretch_column, QHeaderView.ResizeMode.Stretch)
    elif table.columnCount() > 0:
        # Varsayılan olarak son sütun esnek
        header.setSectionResizeMode(table.columnCount() - 1, QHeaderView.ResizeMode.Stretch)
    
    # Minimum genişlik ayarla
    header.setMinimumSectionSize(50)
    
    # Sağ tık menüsü ekle
    header.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
    header.customContextMenuRequested.connect(
        lambda pos: show_column_menu(table, pos, table_id)
    )
    
    # Kayıtlı ayarları yükle
    if table_id:
        load_table_settings(table, table_id)
    
    return table


def show_column_menu(table: QTableWidget, pos, table_id: str = None):
    """Sütun görünürlük menüsü göster"""
    menu = QMenu(table)
    menu.setStyleSheet("""
        QMenu {
            background-color: #ffffff;
            border: 1px solid #e0e0e0;
            border-radius: 8px;
            padding: 5px;
        }
        QMenu::item {
            padding: 8px 20px;
            border-radius: 4px;
        }
        QMenu::item:selected {
            background-color: #e3f2fd;
        }
    """)
    
    # Başlık
    title_action = QAction("📋 Sütunları Göster/Gizle", table)
    title_action.setEnabled(False)
    menu.addAction(title_action)
    menu.addSeparator()
    
    # Her sütun için checkbox
    for col in range(table.columnCount()):
        header = table.horizontalHeaderItem(col)
        col_name = header.text() if header else f"Sütun {col+1}"
        
        action = QAction(col_name, table)
        action.setCheckable(True)
        action.setChecked(not table.isColumnHidden(col))
        action.setData(col)
        action.triggered.connect(lambda checked, c=col: toggle_column(table, c, checked, table_id))
        menu.addAction(action)
    
    menu.addSeparator()
    
    # Tümünü göster
    show_all = QAction("✅ Tümünü Göster", table)
    show_all.triggered.connect(lambda: show_all_columns(table, table_id))
    menu.addAction(show_all)
    
    # Sıfırla
    reset_action = QAction("🔄 Varsayılana Dön", table)
    reset_action.triggered.connect(lambda: reset_table_settings(table, table_id))
    menu.addAction(reset_action)
    
    menu.exec_(table.horizontalHeader().mapToGlobal(pos))


def toggle_column(table: QTableWidget, col: int, visible: bool, table_id: str = None):
    """Sütun görünürlüğünü değiştir"""
    table.setColumnHidden(col, not visible)
    if table_id:
        save_table_settings(table, table_id)


def show_all_columns(table: QTableWidget, table_id: str = None):
    """Tüm sütunları göster"""
    for col in range(table.columnCount()):
        table.setColumnHidden(col, False)
    if table_id:
        save_table_settings(table, table_id)


def save_table_settings(table: QTableWidget, table_id: str):
    """Tablo ayarlarını kaydet"""
    settings = QSettings("BADER", "DernekYonetimi")
    
    # Gizli sütunlar
    hidden_cols = []
    for col in range(table.columnCount()):
        if table.isColumnHidden(col):
            hidden_cols.append(col)
    settings.setValue(f"table/{table_id}/hidden_columns", hidden_cols)
    
    # Sütun genişlikleri
    widths = []
    for col in range(table.columnCount()):
        widths.append(table.columnWidth(col))
    settings.setValue(f"table/{table_id}/column_widths", widths)
    
    # Sütun sıralaması
    order = []
    header = table.horizontalHeader()
    for i in range(table.columnCount()):
        order.append(header.logicalIndex(i))
    settings.setValue(f"table/{table_id}/column_order", order)


def load_table_settings(table: QTableWidget, table_id: str):
    """Tablo ayarlarını yükle"""
    settings = QSettings("BADER", "DernekYonetimi")
    
    # Gizli sütunlar
    hidden_cols = settings.value(f"table/{table_id}/hidden_columns", [])
    if hidden_cols:
        for col in hidden_cols:
            if isinstance(col, int) and col < table.columnCount():
                table.setColumnHidden(col, True)
    
    # Sütun genişlikleri
    widths = settings.value(f"table/{table_id}/column_widths", [])
    if widths:
        for col, width in enumerate(widths):
            if col < table.columnCount() and isinstance(width, int) and width > 0:
                table.setColumnWidth(col, width)


def reset_table_settings(table: QTableWidget, table_id: str):
    """Tablo ayarlarını sıfırla"""
    settings = QSettings("BADER", "DernekYonetimi")
    settings.remove(f"table/{table_id}")
    
    # Tüm sütunları göster
    for col in range(table.columnCount()):
        table.setColumnHidden(col, False)
    
    # Varsayılan boyutlara dön
    header = table.horizontalHeader()
    header.setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)

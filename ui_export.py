"""
BADER Derneği - Export ve Yedekleme İşlemleri
"""

from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QPushButton, 
                             QLabel, QFileDialog, QGroupBox,
                             QTextEdit, QProgressBar)
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from qfluentwidgets import MessageBox
from database import Database
from models import (UyeYoneticisi, AidatYoneticisi, GelirYoneticisi, 
                    GiderYoneticisi, KasaYoneticisi, RaporYoneticisi)
from datetime import datetime
import os


class ExportThread(QThread):
    """Export işlemi için thread"""
    
    progress = pyqtSignal(int, str)
    finished = pyqtSignal(bool, str)
    
    def __init__(self, db: Database, export_type: str, file_path: str):
        super().__init__()
        self.db = db
        self.export_type = export_type
        self.file_path = file_path
        
    def run(self):
        try:
            if self.export_type == "excel":
                self.export_to_excel()
            elif self.export_type == "backup":
                self.backup_database()
            else:
                self.finished.emit(False, "Geçersiz export tipi")
        except Exception as e:
            self.finished.emit(False, str(e))
            
    def export_to_excel(self):
        """Excel dosyasına export"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            self.progress.emit(10, "Workbook oluşturuluyor...")
            
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Varsayılan sheet'i kaldır
            
            # Header stili
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="center")
            
            # 1. Üyeler
            self.progress.emit(20, "Üyeler export ediliyor...")
            uye_yoneticisi = UyeYoneticisi(self.db)
            uyeler = uye_yoneticisi.uye_listesi()
            
            ws_uyeler = wb.create_sheet("Üyeler")
            headers = ["Üye No", "Ad Soyad", "Telefon", "Email", "Durum", "Kayıt Tarihi"]
            ws_uyeler.append(headers)
            
            for col in range(1, len(headers) + 1):
                cell = ws_uyeler.cell(1, col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
            
            for uye in uyeler:
                ws_uyeler.append([
                    uye['uye_id'], uye['ad_soyad'], uye['telefon'], 
                    uye['email'], uye['durum'], uye['kayit_tarihi'][:10]
                ])
            
            # 2. Aidat Takip
            self.progress.emit(35, "Aidat kayıtları export ediliyor...")
            aidat_yoneticisi = AidatYoneticisi(self.db)
            aidatlar = aidat_yoneticisi.aidat_listesi()
            
            ws_aidat = wb.create_sheet("Aidat Takip")
            headers = ["Üye", "Yıl", "Yıllık Aidat", "Toplam Ödenen", "Kalan", "Durum", "Aktarım"]
            ws_aidat.append(headers)
            
            for col in range(1, len(headers) + 1):
                cell = ws_aidat.cell(1, col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
            
            for aidat in aidatlar:
                ws_aidat.append([
                    aidat['ad_soyad'], aidat['yil'], aidat['yillik_aidat_tutari'],
                    aidat['toplam_odenen'], aidat['odenecek_tutar'], 
                    aidat['durum'], aidat['aktarim_durumu']
                ])
            
            # 3. Gelirler
            self.progress.emit(50, "Gelirler export ediliyor...")
            gelir_yoneticisi = GelirYoneticisi(self.db)
            gelirler = gelir_yoneticisi.gelir_listesi()
            
            ws_gelir = wb.create_sheet("Gelirler")
            headers = ["Tarih", "Belge No", "Gelir Türü", "Açıklama", "Tutar", "Kasa", "Tahsil Eden"]
            ws_gelir.append(headers)
            
            for col in range(1, len(headers) + 1):
                cell = ws_gelir.cell(1, col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
            
            for gelir in gelirler:
                ws_gelir.append([
                    gelir['tarih'], gelir['belge_no'], gelir['gelir_turu'],
                    gelir['aciklama'], gelir['tutar'], gelir['kasa_adi'], gelir.get('tahsil_eden', '')
                ])
            
            # 4. Giderler
            self.progress.emit(65, "Giderler export ediliyor...")
            gider_yoneticisi = GiderYoneticisi(self.db)
            giderler = gider_yoneticisi.gider_listesi()
            
            ws_gider = wb.create_sheet("Giderler")
            headers = ["Tarih", "İşlem No", "Gider Türü", "Açıklama", "Tutar", "Kasa", "Ödeyen"]
            ws_gider.append(headers)
            
            for col in range(1, len(headers) + 1):
                cell = ws_gider.cell(1, col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
            
            for gider in giderler:
                ws_gider.append([
                    gider['tarih'], gider['islem_no'], gider['gider_turu'],
                    gider['aciklama'], gider['tutar'], gider['kasa_adi'], gider.get('odeyen', '')
                ])
            
            # 5. Kasa Özeti
            self.progress.emit(80, "Kasa özeti export ediliyor...")
            kasa_yoneticisi = KasaYoneticisi(self.db)
            kasalar = kasa_yoneticisi.tum_kasalar_ozet()
            
            ws_kasa = wb.create_sheet("Kasa Özeti")
            headers = ["Kasa Adı", "Para Birimi", "Devir", "Toplam Gelir", "Toplam Gider", "Net Bakiye"]
            ws_kasa.append(headers)
            
            for col in range(1, len(headers) + 1):
                cell = ws_kasa.cell(1, col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
            
            for kasa in kasalar:
                ws_kasa.append([
                    kasa['kasa_adi'], kasa['para_birimi'], kasa['devir_bakiye'],
                    kasa['toplam_gelir'], kasa['toplam_gider'], kasa['net_bakiye']
                ])
            
            # 6. Genel Rapor
            self.progress.emit(90, "Genel rapor hazırlanıyor...")
            rapor_yoneticisi = RaporYoneticisi(self.db)
            ozet = rapor_yoneticisi.genel_ozet(datetime.now().year)
            
            ws_rapor = wb.create_sheet("Genel Rapor")
            ws_rapor.append(["BADER DERNEĞİ GENEL RAPOR"])
            ws_rapor.append(["Tarih:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
            ws_rapor.append([])
            ws_rapor.append(["Toplam Gelir:", ozet['toplam_gelir']])
            ws_rapor.append(["Toplam Gider:", ozet['toplam_gider']])
            ws_rapor.append(["Net Sonuç:", ozet['net_sonuc']])
            ws_rapor.append(["Toplam Kasa Bakiye:", ozet['toplam_kasa_bakiye']])
            ws_rapor.append([])
            ws_rapor.append(["Toplam Üye:", ozet['toplam_uye']])
            ws_rapor.append(["Aidat Ödeyen Üye:", ozet['aidat_odenen_uye']])
            ws_rapor.append(["Aidat Eksik Üye:", ozet['aidat_eksik_uye']])
            
            # Kaydet
            self.progress.emit(95, "Dosya kaydediliyor...")
            wb.save(self.file_path)
            
            self.progress.emit(100, "Tamamlandı!")
            self.finished.emit(True, "Excel dosyası başarıyla oluşturuldu!")
            
        except ImportError:
            self.finished.emit(False, "openpyxl kütüphanesi bulunamadı! 'pip install openpyxl' komutunu çalıştırın.")
        except Exception as e:
            self.finished.emit(False, f"Export hatası: {str(e)}")
            
    def backup_database(self):
        """Veritabanını yedekle"""
        self.progress.emit(50, "Yedekleme yapılıyor...")
        
        success = self.db.backup_database(self.file_path)
        
        self.progress.emit(100, "Tamamlandı!")
        
        if success:
            self.finished.emit(True, "Veritabanı başarıyla yedeklendi!")
        else:
            self.finished.emit(False, "Yedekleme sırasında hata oluştu!")


class ExportWidget(QWidget):
    """Export ve yedekleme widget"""
    
    def __init__(self, db: Database):
        super().__init__()
        self.db = db
        self.export_thread = None
        self.setup_ui()
        
    def setup_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(15)
        
        # Başlık
        title_label = QLabel("EXPORT & YEDEKLEME")
        title_label.setProperty("class", "title")
        layout.addWidget(title_label)
        
        # Excel Export
        excel_group = QGroupBox("Excel Export")
        excel_layout = QVBoxLayout()
        
        excel_info = QLabel(
            "📊 Tüm verileri Excel dosyasına export edin.\n"
            "İçerik: Üyeler, Aidat, Gelirler, Giderler, Kasa Özeti, Genel Rapor"
        )
        excel_info.setWordWrap(True)
        excel_layout.addWidget(excel_info)
        
        self.excel_btn = QPushButton("📥 Excel'e Export Et")
        self.excel_btn.clicked.connect(self.export_excel)
        excel_layout.addWidget(self.excel_btn)
        
        excel_group.setLayout(excel_layout)
        layout.addWidget(excel_group)
        
        # Veritabanı Yedekleme
        backup_group = QGroupBox("Veritabanı Yedekleme")
        backup_layout = QVBoxLayout()
        
        backup_info = QLabel(
            "💾 Tüm veritabanını yedekleyin.\n"
            "Yedeklenen dosyayı güvenli bir yerde saklayın."
        )
        backup_info.setWordWrap(True)
        backup_layout.addWidget(backup_info)
        
        backup_btn_layout = QHBoxLayout()
        
        self.backup_btn = QPushButton("💾 Yedekle")
        self.backup_btn.clicked.connect(self.backup_db)
        backup_btn_layout.addWidget(self.backup_btn)
        
        self.restore_btn = QPushButton("📂 Geri Yükle")
        self.restore_btn.setProperty("class", "warning")
        self.restore_btn.clicked.connect(self.restore_db)
        backup_btn_layout.addWidget(self.restore_btn)
        
        backup_layout.addLayout(backup_btn_layout)
        
        backup_group.setLayout(backup_layout)
        layout.addWidget(backup_group)
        
        # Progress
        progress_group = QGroupBox("İşlem Durumu")
        progress_layout = QVBoxLayout()
        
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        progress_layout.addWidget(self.progress_bar)
        
        self.progress_label = QLabel("")
        self.progress_label.setVisible(False)
        progress_layout.addWidget(self.progress_label)
        
        progress_group.setLayout(progress_layout)
        layout.addWidget(progress_group)
        
        # Log
        log_group = QGroupBox("İşlem Geçmişi")
        log_layout = QVBoxLayout()
        
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setMaximumHeight(200)
        log_layout.addWidget(self.log_text)
        
        log_group.setLayout(log_layout)
        layout.addWidget(log_group)
        
        layout.addStretch()
        
        self.setLayout(layout)
        
    def export_excel(self):
        """Excel'e export"""
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Excel Dosyası Kaydet",
            f"BADER_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel Files (*.xlsx)"
        )
        
        if file_path:
            self.start_export("excel", file_path)
            
    def backup_db(self):
        """Veritabanını yedekle"""
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Yedekleme Dosyası Kaydet",
            f"BADER_Backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db",
            "Database Files (*.db)"
        )
        
        if file_path:
            self.start_export("backup", file_path)
            
    def restore_db(self):
        """Veritabanını geri yükle"""
        w = MessageBox("Uyarı", 
                      "⚠️ Mevcut veritabanı yerine yedekleme dosyası yüklenecektir!\n"
                      "Bu işlem geri alınamaz. Devam etmek istiyor musunuz?",
                      self)
        if w.exec():
            file_path, _ = QFileDialog.getOpenFileName(
                self,
                "Yedekleme Dosyası Seç",
                "",
                "Database Files (*.db)"
            )
            
            if file_path:
                try:
                    success = self.db.restore_database(file_path)
                    if success:
                        MessageBox("Başarılı", "Veritabanı başarıyla geri yüklendi!\n"
                            "Lütfen programı yeniden başlatın."
                        , 
                            self).show()
                        self.add_log("Veritabanı geri yüklendi: " + file_path)
                    else:
                        MessageBox("Hata", "Geri yükleme başarısız!", self).show()
                except Exception as e:
                    MessageBox("Hata", f"Geri yükleme hatası:\n{e}", self).show()
                    
    def start_export(self, export_type: str, file_path: str):
        """Export işlemini başlat"""
        self.progress_bar.setVisible(True)
        self.progress_label.setVisible(True)
        self.progress_bar.setValue(0)
        
        self.excel_btn.setEnabled(False)
        self.backup_btn.setEnabled(False)
        self.restore_btn.setEnabled(False)
        
        self.export_thread = ExportThread(self.db, export_type, file_path)
        self.export_thread.progress.connect(self.on_progress)
        self.export_thread.finished.connect(self.on_finished)
        self.export_thread.start()
        
    def on_progress(self, value: int, message: str):
        """Progress güncelle"""
        self.progress_bar.setValue(value)
        self.progress_label.setText(message)
        
    def on_finished(self, success: bool, message: str):
        """Export tamamlandı"""
        self.excel_btn.setEnabled(True)
        self.backup_btn.setEnabled(True)
        self.restore_btn.setEnabled(True)
        
        if success:
            MessageBox("Başarılı", message, self).show()
            self.add_log("✓ " + message)
        else:
            MessageBox("Hata", message, self).show()
            self.add_log("✗ " + message)
            
        self.progress_bar.setVisible(False)
        self.progress_label.setVisible(False)
        
    def add_log(self, message: str):
        """Log ekle"""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.log_text.append(f"[{timestamp}] {message}")


